import sys
import json
import os
import pyodbc
from openpyxl import load_workbook
import shutil
from datetime import datetime

sys.path.append('C:\\Temp')

import Automation_Tools


def SheetLoaderMain():

    for xlsFileName in xlsFilesArray:
        
        objWBook = load_workbook(ServerFolder + xlsFileName)

        if (ImportAllSheets == True):
            
            for objSheet in objWBook.get_sheet_names():
                ReadSheetData(objWBook.get_sheet_by_name(objSheet), xlsFileName)
                
        else:

            ReadSheetData(objWBook.worksheets[0], xlsFileName)

        objWBook.close()

        if (MoveAfterProcessed):
            shutil.move(ServerFolder + xlsFileName, ProcessedFolder +  xlsFileName)

def HeaderCheckedOk():
    return True

def ReadSheetData(objSheet, xlsFileName):
    
    if (HeaderCheckedOk()):
        lngLastRow = objSheet.max_row       
        lngCurrentRow = FirstDataRow

        objConn = pyodbc.connect(SanofiReportsSQLConnectionString, autocommit=False)
        objCursor = objConn.cursor()

        while (lngCurrentRow <= lngLastRow):

            valores = []
            strDebugString = ''

            for DataField in FieldMap: 

                strDataType = DataField['Field'][3].upper()
                intDataColumn = DataField['Field'][0]
                currentValue = objSheet.cell(row=lngCurrentRow, column=intDataColumn).value
                
                if strDataType == "HR":
                    #Converte objeto Hora para texto
                    if(currentValue != '' and currentValue != None):
                        currentValue = currentValue.strftime('%H:%M:%S')
                elif strDataType == "DT":
                    #Converte objeto data para texto
                    
                    if(currentValue != '' and currentValue != None):
                        
                        if (type(currentValue) is datetime.date):                        
                            currentValue = currentValue.strftime("%Y-%d-%m")

                        elif (isinstance(currentValue, str)):
                            currentValue = currentValue[6:10] + "-" + currentValue[3:5] + "-" + currentValue[0:2]

                elif (strDataType in ("MN","NR","FL")):
                    #Converte objeto data para texto
                    if (currentValue is None):
                        currentValue = 0                   
                    elif (type(currentValue) is str):                       
                        currentValue = currentValue.replace(".","")
                        currentValue = currentValue.replace(",",".")
                        try:
                            int(currentValue)
                        except:
                            currentValue = 0

                valores.append(currentValue)

                strDebugString = strDebugString + DataField['Field'][1] + ' = "' + str(currentValue) + '"; '
            
            try:

                objCursor.execute('EXEC ' + StoredProcedure + storedProcedureParameters, (valores))              #Executa a chamada da Stored Procedure no banco
                objConn.commit()
                print(lngCurrentRow)      

            except pyodbc.Error as pyodbcError:

                id_evento = Automation_Tools.CreateGUID()
                dt_evento = Automation_Tools.NowDateTime()
                vl_tipo_evento = intExecutionType 
                vl_execucao = intExecutionType

                st_evento = ('Stored Procedure: \n' + StoredProcedure + '\n' +
                                'Pasta Origem: \n' + ServerFolder + '\n' +
                                'Arquivo Excel: \n' + xlsFileName + '\n' +
                                'Linha: \n' + str(lngCurrentRow) + '\n' +
                                'Erro: \n' + pyodbcError.args[1] + '\n' +
                                'Dados: \n"' + strDebugString + '\n' +
                                'Parâmetros: \n' + storedProcedureParameters)

                #Em caso de erro grava um registro na base de erros de automação
                    
                Automation_Tools.eventos_automacao(id_batch, 
                                                    id_evento, 
                                                    dt_evento, 
                                                    ProcessName, 
                                                    vl_tipo_evento, 
                                                    st_evento, 
                                                    vl_execucao)
                
                print(str(lngCurrentRow))

            lngCurrentRow += 1

        objCursor.close()
        objConn.close()
        
    else:
        print("Header Not Ok")

intExecutionType = 0
JSONFile = Automation_Tools.GetSheetLoaderFile("SAPZBRSD0286")
#JSONFile = Automation_Tools.GetSheetLoaderFile("DeParaTransportadoras")

with open(JSONFile, "r",  encoding="utf8") as SheetLoaderJSON:
    data = json.load(SheetLoaderJSON)

    ProcessName = data['ProcessName']
    ServerFolder = data['ServerFolder']
    ProcessedFolder = data['ProcessedFolder']
    MoveAfterProcessed = data['MoveAfterProcessed']
    MoveOnError = data['MoveOnError']
    FileNameWildCard = data['FileNameWildCard']
    AcceptedExtensions = data['AcceptedExtensions']
    FirstDataRow = data['FirstDataRow']
    StoredProcedure = data['StoredProcedure']
    CheckSheetHeader = data['CheckSheetHeader']
    ImportAllSheets = data['ImportAllSheets']

    FieldMap = data['FieldMap']

storedProcedureParameters = ''

xlsFilesArray = [fn for fn in os.listdir(ServerFolder) if fn.split(".")[-1].lower() in AcceptedExtensions and fn[0:len(FileNameWildCard)] == FileNameWildCard]

if (len(xlsFilesArray) > 0):

    id_batch = Automation_Tools.CreateGUID()

    for DataField in FieldMap:        
        storedProcedureParameters = storedProcedureParameters + ' @' + DataField['Field'][2] + ' = ? , '

    storedProcedureParameters = storedProcedureParameters[:-3]

    SanofiReportsSQLConnectionString = Automation_Tools.strSanofiReportsConnectionString

    SheetLoaderMain()

else:

    print('Nenhum arquivo excel foi encontrado.')